"""Shared helpers for the collection scripts.

Handles Google API auth, xlsx loading, and the canonical row schema.
"""
from __future__ import annotations

import os
import json
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl import Workbook

SCOPES_GMAIL = ["https://www.googleapis.com/auth/gmail.readonly"]
SCOPES_CALENDAR = ["https://www.googleapis.com/auth/calendar.readonly"]

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_XLSX = REPO_ROOT / "yoga_visits.xlsx"
TEMPLATE_XLSX = REPO_ROOT / "templates" / "yoga_visits_template.xlsx"
CREDENTIALS = REPO_ROOT / "credentials.json"
TOKEN_GMAIL = REPO_ROOT / "token_gmail.json"
TOKEN_CALENDAR = REPO_ROOT / "token_calendar.json"

COLUMNS = [
    "Date", "Day", "Time", "Style (Harmonized)", "Class Name (Raw)",
    "Teacher", "Studio", "City", "State/Province", "Country",
    "Source", "Notes", "Unsure Attended", "Studio Group",
    "Check-in Confirmed", "Swarm Venue", "Studio Status",
    "Eventbrite Confirmed", "Training",
]


def get_gmail_service():
    """Return an authorized Gmail API service object."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if TOKEN_GMAIL.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_GMAIL), SCOPES_GMAIL)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS), SCOPES_GMAIL)
            creds = flow.run_local_server(port=0)
        TOKEN_GMAIL.write_text(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def get_calendar_service():
    """Return an authorized Calendar API service object."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if TOKEN_CALENDAR.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_CALENDAR), SCOPES_CALENDAR)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS), SCOPES_CALENDAR)
            creds = flow.run_local_server(port=0)
        TOKEN_CALENDAR.write_text(creds.to_json())
    return build("calendar", "v3", credentials=creds)


def load_or_create_workbook(path: Path) -> openpyxl.Workbook:
    """Open the xlsx, creating from template (or empty) if missing."""
    if path.exists():
        return openpyxl.load_workbook(path)
    if TEMPLATE_XLSX.exists():
        wb = openpyxl.load_workbook(TEMPLATE_XLSX)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Yoga Visits"
        for i, col in enumerate(COLUMNS, 1):
            ws.cell(1, i, col)
    return wb


def to_date(d) -> Optional[date]:
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    if isinstance(d, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y"):
            try:
                return datetime.strptime(d, fmt).date()
            except ValueError:
                continue
    return None


def find_matching_row(ws, the_date: date, studio: str, time_str: str = None) -> Optional[int]:
    """Return the row index of an existing matching visit, or None.

    Match key: date + studio (+ time when provided).
    """
    for r in range(2, ws.max_row + 1):
        d = to_date(ws.cell(r, 1).value)
        s = ws.cell(r, 7).value
        if d == the_date and (s or "").strip().lower() == (studio or "").strip().lower():
            if time_str is None:
                return r
            t = ws.cell(r, 3).value
            if (t or "").strip() == time_str.strip():
                return r
    return None


def append_row(ws, row_dict: dict) -> int:
    """Append a new row built from a dict keyed by column names."""
    r = ws.max_row + 1
    for i, col in enumerate(COLUMNS, 1):
        ws.cell(r, i, row_dict.get(col))
    # Derive Day if missing
    if not row_dict.get("Day"):
        d = to_date(row_dict.get("Date"))
        if d:
            ws.cell(r, 2, d.strftime("%A"))
    return r


def stamp_row(ws, r: int, updates: dict, source_tag: str = None):
    """Update an existing row with new fields. Source tag is appended."""
    for col, val in updates.items():
        if col == "Source":
            continue
        if col in COLUMNS:
            idx = COLUMNS.index(col) + 1
            if val and not ws.cell(r, idx).value:
                ws.cell(r, idx, val)
    if source_tag:
        existing = (ws.cell(r, 11).value or "").strip()
        if source_tag not in existing:
            ws.cell(r, 11, f"{existing} + {source_tag}" if existing else source_tag)


def upsert_visit(ws, row_dict: dict, strategy: str = "stamp") -> str:
    """Insert or update a visit row.

    Returns: 'inserted', 'stamped', or 'skipped'
    """
    d = to_date(row_dict.get("Date"))
    if not d:
        return "skipped"
    existing = find_matching_row(ws, d, row_dict.get("Studio") or "", row_dict.get("Time"))
    if existing is None:
        append_row(ws, row_dict)
        return "inserted"
    if strategy == "stamp":
        stamp_row(ws, existing, row_dict, source_tag=row_dict.get("Source"))
        return "stamped"
    if strategy == "append":
        append_row(ws, row_dict)
        return "inserted"
    return "skipped"
