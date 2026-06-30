"""Normalize teacher names, studio names, and styles in yoga_visits.xlsx.

Idempotent — safe to re-run. Reads aliases from
config/teacher_aliases.yaml and config/studio_aliases.yaml.

Usage:
    python scripts/analyze/harmonize.py
    python scripts/analyze/harmonize.py --input yoga_visits.xlsx --dry-run
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = REPO_ROOT / "yoga_visits.xlsx"
TEACHER_ALIASES = REPO_ROOT / "config" / "teacher_aliases.yaml"
STUDIO_ALIASES = REPO_ROOT / "config" / "studio_aliases.yaml"
STYLE_BUCKETS = REPO_ROOT / "config" / "style_buckets.yaml"


def build_reverse_map(alias_doc):
    """Given {canonical: [alias1, alias2]}, return {alias: canonical}."""
    reverse = {}
    for canonical, aliases in (alias_doc.get("aliases") or {}).items():
        for a in aliases or []:
            reverse[a.strip().lower()] = canonical
    return reverse


def build_style_map(style_doc):
    """Given the style_buckets.yaml structure, return {alias: bucket}."""
    reverse = {}
    for bucket, body in (style_doc.get("buckets") or {}).items():
        for a in (body.get("aliases") or []):
            reverse[a.strip().lower()] = bucket
    return reverse


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", default=str(DEFAULT_INPUT))
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    in_path = Path(args.input)
    wb = openpyxl.load_workbook(in_path)
    ws = wb["Yoga Visits"]

    teacher_map = build_reverse_map(_load_yaml(TEACHER_ALIASES))
    studio_map = build_reverse_map(_load_yaml(STUDIO_ALIASES))
    style_map = build_style_map(_load_yaml(STYLE_BUCKETS))

    changes = {"teacher": 0, "studio": 0, "style": 0, "country": 0}
    country_norms = {
        "united states": "USA", "united states of america": "USA", "u.s.": "USA", "us": "USA",
        "united kingdom": "UK", "u.k.": "UK",
    }

    for r in range(2, ws.max_row + 1):
        # Teacher (col 6) — handle " + " co-teachers
        t = ws.cell(r, 6).value
        if t:
            parts = [p.strip() for p in t.split(" + ")]
            new_parts = [teacher_map.get(p.lower(), p) for p in parts]
            new_t = " + ".join(new_parts)
            if new_t != t:
                if not args.dry_run:
                    ws.cell(r, 6, new_t)
                changes["teacher"] += 1

        # Studio (col 7)
        s = ws.cell(r, 7).value
        if s and s.lower() in studio_map:
            new_s = studio_map[s.lower()]
            if not args.dry_run:
                ws.cell(r, 7, new_s)
            changes["studio"] += 1

        # Style (col 4)
        st = ws.cell(r, 4).value
        if st and st.lower() in style_map:
            new_st = style_map[st.lower()]
            if new_st != st:
                if not args.dry_run:
                    ws.cell(r, 4, new_st)
                changes["style"] += 1

        # Country (col 10)
        c = ws.cell(r, 10).value
        if c and c.lower() in country_norms:
            new_c = country_norms[c.lower()]
            if new_c != c:
                if not args.dry_run:
                    ws.cell(r, 10, new_c)
                changes["country"] += 1

    if not args.dry_run:
        wb.save(in_path)

    print(f"{'[dry-run] ' if args.dry_run else ''}Changes: {changes}")


def _load_yaml(p: Path):
    if not p.exists():
        return {}
    with open(p) as f:
        return yaml.safe_load(f) or {}


if __name__ == "__main__":
    main()
