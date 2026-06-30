"""Set the Training tag (col 19) based on class name, studio, and notes.

Idempotent — existing tags are preserved unless --overwrite is passed.

Usage:
    python scripts/analyze/classify_training.py
    python scripts/analyze/classify_training.py --overwrite
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = REPO_ROOT / "yoga_visits.xlsx"
TAGS_CONFIG = REPO_ROOT / "config" / "training_tags.yaml"

# Heuristic patterns for inferring training tags from raw class/studio/notes text
PATTERNS = [
    (r"200hr|200 hour", "200hr YTT"),
    (r"300hr|300 hour", "300hr (in progress)"),
    (r"500hr|500 hour", "500hr YTT"),
    (r"aerial.*ytt|aerial.*teacher train", "50hr (Aerial)"),
    (r"restorative.*ytt|restorative.*module", "25hr (Restorative)"),
    (r"anatomy.*ytt|anatomy.*module|anatomy intensive", "25hr (Anatomy)"),
    (r"inversion.*yacep|inversions intensive|awakening inversions", "25hr (Inversions)"),
    (r"sadhana retreat|sadhana intensive", "25hr (Sadhana)"),
    (r"philosophy.*module|philosophy intensive", "25hr (Philosophy)"),
    (r"breath.*workshop|pranayama workshop", "Breath workshop"),
    (r"mantra workshop|chant workshop", "Mantra workshop"),
    (r"workshop|masterclass|clinic", "Workshop"),
    (r"day retreat|one day retreat|day immersion", "Day retreat"),
    (r"retreat", "Multi-day retreat"),
]


def infer_tag(class_name, studio, notes):
    text = " ".join([class_name or "", studio or "", notes or ""]).lower()
    for pat, tag in PATTERNS:
        if re.search(pat, text):
            return tag
    return ""


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", default=str(DEFAULT_INPUT))
    p.add_argument("--overwrite", action="store_true",
                   help="Replace existing tags instead of preserving them")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    in_path = Path(args.input)
    wb = openpyxl.load_workbook(in_path)
    ws = wb["Yoga Visits"]

    set_count = 0
    skipped = 0
    for r in range(2, ws.max_row + 1):
        existing = (ws.cell(r, 19).value or "").strip()
        if existing and not args.overwrite:
            skipped += 1
            continue
        class_name = ws.cell(r, 5).value
        studio = ws.cell(r, 7).value
        notes = ws.cell(r, 12).value
        tag = infer_tag(class_name, studio, notes)
        if tag:
            if not args.dry_run:
                ws.cell(r, 19, tag)
            set_count += 1

    if not args.dry_run:
        wb.save(in_path)

    print(f"{'[dry-run] ' if args.dry_run else ''}Set {set_count} tags, preserved {skipped} existing")


if __name__ == "__main__":
    main()
