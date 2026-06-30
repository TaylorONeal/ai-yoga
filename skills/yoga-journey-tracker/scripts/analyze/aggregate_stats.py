"""Aggregate yoga_visits.xlsx into a stats JSON file.

Produces the canonical input for the output renderers (deck, dashboard,
teachers doc). Enforces the three-bucket model: formal training hours,
workshop hours, and retreat sessions (not hours).

Usage:
    python scripts/analyze/aggregate_stats.py
    python scripts/analyze/aggregate_stats.py --input yoga_visits.xlsx --output stats.json
"""
from __future__ import annotations

import argparse
import json
from datetime import datetime, date
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = REPO_ROOT / "yoga_visits.xlsx"
DEFAULT_OUTPUT = REPO_ROOT / "stats.json"
TAGS_CONFIG = REPO_ROOT / "config" / "training_tags.yaml"
PRACTITIONER_CONFIG = REPO_ROOT / "config" / "practitioner.yaml"


def to_date(d):
    if isinstance(d, datetime):
        return d.date()
    if isinstance(d, date):
        return d
    if isinstance(d, str):
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%B %d, %Y"):
            try:
                return datetime.strptime(d, fmt).date()
            except ValueError:
                continue
    return None


def load_rows(ws):
    rows = []
    for r in range(2, ws.max_row + 1):
        d = to_date(ws.cell(r, 1).value)
        if not d:
            continue
        rows.append({
            "date": d,
            "year": d.year,
            "month": d.strftime("%Y-%m"),
            "dow": d.strftime("%A"),
            "time": ws.cell(r, 3).value,
            "style": ws.cell(r, 4).value or "Other",
            "class_name": ws.cell(r, 5).value or "",
            "teacher": ws.cell(r, 6).value or "",
            "studio": ws.cell(r, 7).value or "",
            "city": ws.cell(r, 8).value or "",
            "country": ws.cell(r, 10).value or "",
            "source": ws.cell(r, 11).value or "",
            "notes": ws.cell(r, 12).value or "",
            "studio_group": ws.cell(r, 14).value or "",
            "training": (ws.cell(r, 19).value or "").strip(),
        })
    return rows


def compute_training_hours(rows, tags_config, practitioner_config):
    """Three-bucket sum.

    Formal hours come from `featured_trainings` in practitioner config
    if present, else inferred from tag counts × estimated hours per tag.
    Workshops are tag-counted × 2.5 hr average.
    Retreats are session-counted (NOT hours).
    """
    formal_tags = set(tags_config.get("training_tags", []))
    workshop_tags = set(tags_config.get("workshop_tags", []))
    retreat_tags = set(tags_config.get("retreat_tags", []))

    tag_counts = Counter(r["training"] for r in rows if r["training"])

    # Formal hours: prefer practitioner-config explicit values
    formal_hours = 0
    in_progress_hours = 0
    featured = practitioner_config.get("featured_trainings", []) if practitioner_config else []
    in_progress = practitioner_config.get("in_progress_trainings", []) if practitioner_config else []

    if featured:
        formal_hours = sum(t.get("hours", 0) for t in featured)
    else:
        # Fallback: assume each formal-tag is a complete training (rough)
        hour_estimates = {
            "200hr YTT": 200, "300hr YTT": 300, "500hr YTT": 500,
            "50hr (Aerial)": 50, "50hr (Yin)": 50, "50hr (Restorative)": 50,
            "25hr (Restorative)": 25, "25hr (Anatomy)": 25, "25hr (Inversions)": 25,
            "25hr (Philosophy)": 25, "25hr (Sadhana)": 25,
        }
        for tag, count in tag_counts.items():
            if tag in formal_tags and tag in hour_estimates:
                # Tag appears once per training, not per session
                formal_hours += hour_estimates[tag]

    if in_progress:
        in_progress_hours = sum(t.get("hours", 0) for t in in_progress)
    elif "300hr (in progress)" in tag_counts:
        in_progress_hours = 300

    # Workshop hours: rough estimate (2.5 hr/workshop)
    workshop_sessions = sum(c for t, c in tag_counts.items() if t in workshop_tags)
    workshop_hours = round(workshop_sessions * 2.5)

    # Retreat session count
    retreat_sessions = sum(c for t, c in tag_counts.items() if t in retreat_tags)

    return {
        "formal": formal_hours,
        "in_progress": in_progress_hours,
        "workshops": workshop_hours,
        "workshop_sessions": workshop_sessions,
        "retreat_sessions": retreat_sessions,
    }


def sanity_checks(rows):
    """Emit warnings for likely data-quality issues."""
    warnings = []
    blank_country = sum(1 for r in rows if not r["country"])
    if blank_country:
        warnings.append(f"{blank_country} rows missing Country")

    # Day-of-week mismatch with date
    mismatches = sum(1 for r in rows if False)  # placeholder; Day is derived
    if mismatches:
        warnings.append(f"{mismatches} rows have Day not matching Date")

    # Year gaps
    years = sorted({r["year"] for r in rows})
    if len(years) > 2:
        for i in range(1, len(years) - 1):
            if years[i + 1] - years[i] > 1:
                warnings.append(f"Year gap between {years[i]} and {years[i + 1]}")

    return warnings


def main():
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", default=str(DEFAULT_INPUT))
    p.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = p.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb["Yoga Visits"]
    rows = load_rows(ws)

    # Load configs
    tags_config = {}
    if TAGS_CONFIG.exists():
        with open(TAGS_CONFIG) as f:
            tags_config = yaml.safe_load(f) or {}

    practitioner_config = {}
    if PRACTITIONER_CONFIG.exists():
        with open(PRACTITIONER_CONFIG) as f:
            practitioner_config = yaml.safe_load(f) or {}

    # Core aggregations
    years = Counter(r["year"] for r in rows)
    months = Counter(r["month"] for r in rows)
    styles = Counter(r["style"] for r in rows)
    teachers = Counter(r["teacher"] for r in rows if r["teacher"])
    studios = Counter(r["studio"] for r in rows if r["studio"])
    countries = Counter(r["country"] for r in rows if r["country"])

    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    dow_counts = Counter(r["dow"] for r in rows)
    dow_sorted = [(d, dow_counts.get(d, 0)) for d in days_order]

    training_hours = compute_training_hours(rows, tags_config, practitioner_config)
    training_tag_counts = Counter(r["training"] for r in rows if r["training"])

    intl_visits = sum(1 for r in rows if r["country"] and r["country"] != "USA")
    first_year = min(years.keys())
    last_year = max(years.keys())
    years_active = last_year - first_year + 1

    total = len(rows)

    data = {
        "total_visits": total,
        "first_year": first_year,
        "last_year": last_year,
        "years_active": years_active,
        "unique_studios": len(studios),
        "unique_teachers": len(teachers),
        "unique_countries": len(countries),
        "unique_cities": len({r["city"] for r in rows if r["city"]}),
        "avg_per_year": round(total / years_active, 1),
        "yearly": [{"year": y, "count": years[y]} for y in sorted(years)],
        "monthly": [{"month": m, "count": months[m]} for m in sorted(months)],
        "top_studios": [{"studio": s, "count": c} for s, c in studios.most_common(15)],
        "top_teachers": [{"teacher": t, "count": c} for t, c in teachers.most_common(15)],
        "top_countries": [{"country": c, "count": n} for c, n in countries.most_common(10)],
        "styles": [
            {"style": s, "count": c, "pct": round(c / total * 100, 1)}
            for s, c in styles.most_common()
        ],
        "dow": [{"day": d, "count": c} for d, c in dow_sorted],
        "training_tags": [{"tag": t, "count": c} for t, c in training_tag_counts.most_common()],
        "training_hours": training_hours,
        "retreats": {
            "session_count": training_hours["retreat_sessions"],
        },
        "peak_year": list(max(years.items(), key=lambda x: x[1])),
        "dna": {
            "intl_visits": intl_visits,
            "lifetime_avg_per_week": round(total / (years_active * 52), 1),
        },
        "practitioner": practitioner_config,
    }

    with open(out_path, "w") as f:
        json.dump(data, f, indent=2, default=str)

    print(f"Wrote {out_path}")
    print(f"  {total} visits  ·  {years_active} years  ·  {len(studios)} studios  ·  {len(teachers)} teachers")
    print(f"  Training: {training_hours['formal']}h formal · {training_hours['workshops']}h workshops · {training_hours['retreat_sessions']} retreat sessions")
    print(f"  In progress: {training_hours['in_progress']}h")

    warnings = sanity_checks(rows)
    if warnings:
        print("\nSanity warnings:")
        for w in warnings:
            print(f"  · {w}")


if __name__ == "__main__":
    main()
